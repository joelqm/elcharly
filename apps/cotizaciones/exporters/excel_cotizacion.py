"""Exportación Excel de cotización (formato STA editable)."""
from __future__ import annotations

from io import BytesIO
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.cotizaciones.models import DOCUMENTO_VERSION


def exportar_cotizacion_excel(cotizacion) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cotización'

    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    header_fill = PatternFill('solid', fgColor='0F172A')
    label_fill = PatternFill('solid', fgColor='F1F5F9')
    total_fill = PatternFill('solid', fgColor='ECFDF5')
    header_font = Font(bold=True, name='Calibri', size=10, color='FFFFFF')
    bold = Font(bold=True, name='Calibri', size=11, color='0F172A')
    title_font = Font(bold=True, name='Calibri', size=18, color='0F172A')
    script_font = Font(bold=True, name='Calibri', size=14, color='009B94', italic=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Logo
    logo_path = Path(settings.MEDIA_ROOT) / 'logo_makita.png'
    if not logo_path.is_file():
        logo_path = Path(settings.BASE_DIR) / 'media' / 'logo_makita.png'
    if logo_path.is_file():
        try:
            img = XLImage(str(logo_path))
            img.width = 120
            img.height = 44
            ws.add_image(img, 'A1')
        except Exception:
            ws['A1'] = 'Makita'
            ws['A1'].font = Font(bold=True, color='C41E3A', size=12)
    else:
        ws['A1'] = 'Makita'
        ws['A1'].font = Font(bold=True, color='C41E3A', size=12)

    ws.merge_cells('C1:D1')
    ws['C1'] = 'COTIZACIÓN'
    ws['C1'].font = title_font
    ws['C1'].alignment = center

    ws.merge_cells('C2:D2')
    ws['C2'] = 'SERVICIO TÉCNICO AUTORIZADO'
    ws['C2'].alignment = center
    ws['C2'].font = Font(bold=True, size=9, color='64748B')

    ws.merge_cells('C3:D3')
    ws['C3'] = 'El Charly'
    ws['C3'].font = script_font
    ws['C3'].alignment = center

    # Meta alineada a columnas E-F (mismo ancho visual que tabla A-F)
    ws['E1'] = 'CÓDIGO'
    ws['F1'] = cotizacion.numero
    ws['E2'] = 'VERSIÓN'
    ws['F2'] = DOCUMENTO_VERSION
    ws['E3'] = 'FECHA'
    ws['F3'] = cotizacion.fecha_creacion.strftime('%d/%m/%Y')
    for cell in ('E1', 'E2', 'E3'):
        ws[cell].font = Font(bold=True, size=9, color='64748B')
        ws[cell].border = thin
        ws[cell].fill = label_fill
    for cell in ('F1', 'F2', 'F3'):
        ws[cell].border = thin
        ws[cell].alignment = center
        ws[cell].font = bold

    # Cliente ancho completo A-F
    row = 5
    cliente_rows = [
        ('RUC DE CLIENTE', cotizacion.cliente_ruc),
        ('CLIENTE', cotizacion.cliente_nombre),
        ('DIRECCIÓN', cotizacion.cliente_direccion),
        ('N° CELULAR', cotizacion.cliente_telefono),
    ]
    for label, val in cliente_rows:
        ws.cell(row, 1, label).font = Font(bold=True, size=9, color='64748B')
        ws.cell(row, 1).border = thin
        ws.cell(row, 1).fill = label_fill
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = ws.cell(row, 2, val or '')
        c.border = thin
        c.font = bold if label == 'CLIENTE' else Font(name='Calibri', size=11)
        for col in range(2, 7):
            ws.cell(row, col).border = thin
        row += 1

    # Tabla ítems A-F
    row += 1
    headers = ['ÍTEM', 'CÓDIGO', 'DESCRIPCIÓN', 'PRECIO UNITARIO', 'CANT.', 'VALOR']
    start_table = row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = center

    row += 1
    detalles = list(cotizacion.detalles.all().select_related('repuesto'))
    for i, d in enumerate(detalles, start=1):
        vals = [
            i,
            d.codigo_linea,
            d.descripcion_linea,
            f'S/ {Decimal(d.precio_unitario):,.2f}',
            d.cantidad,
            f'S/ {Decimal(d.subtotal):,.2f}',
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row, col, v)
            cell.border = thin
            cell.alignment = center if col != 3 else left
        row += 1

    while row < start_table + 1 + 12:
        for col in range(1, 7):
            ws.cell(row, col, '').border = thin
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    tlab = ws.cell(row, 1, 'TOTAL')
    tlab.font = bold
    tlab.alignment = Alignment(horizontal='right', vertical='center')
    for col in range(1, 5):
        ws.cell(row, col).border = thin
        ws.cell(row, col).fill = total_fill
    tot = ws.cell(row, 5, '')
    tot.border = thin
    tot.fill = total_fill
    tot2 = ws.cell(row, 6, f'S/ {Decimal(cotizacion.total):,.2f}')
    tot2.font = bold
    tot2.border = thin
    tot2.alignment = center
    tot2.fill = total_fill

    row += 2
    for line in cotizacion.lineas_observaciones():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, line)
        cell.alignment = center
        cell.font = bold if row == start_table + 15 else Font(name='Calibri', size=10, color='334155')
        row += 1

    ws2 = wb.create_sheet('Interno (no entregar)')
    ws2.append(['ÍTEM', 'CÓDIGO', 'DESCRIPCIÓN', 'COSTO', 'LISTA SIN IGV', 'P. VENTA C/IGV', 'CANT', 'VALOR', 'MARGEN %'])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = label_fill
    for i, d in enumerate(detalles, start=1):
        costo = Decimal(d.precio_costo or 0)
        lista = Decimal(d.precio_lista or 0)
        venta = Decimal(d.precio_unitario or 0)
        margen = ''
        if costo > 0 and venta > 0:
            margen = round(float((venta - costo) / venta * 100), 1)
        ws2.append([
            i, d.codigo_linea, d.descripcion_linea,
            float(costo), float(lista), float(venta),
            d.cantidad, float(d.subtotal), margen,
        ])

    widths = [8, 14, 48, 16, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
