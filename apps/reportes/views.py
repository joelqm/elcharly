from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Count, F, Q
from django.utils import timezone
from django.http import HttpResponse
from decimal import Decimal
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.pedidos.models import Pedido, DetallePedido
from apps.tienda.models import Producto
from apps.mantenimiento.models import EquipoRegistrado
from apps.cotizaciones.models import Cotizacion

def reportes_required(view_func):
    """
    Decorator to restrict access to Admins, Sellers, and Technicians.
    """
    def _wrapped_view(request, *args, **kwargs):
        from apps.sistema.internal_access import (
            is_staff_interno,
            ocultar_sistema_interno,
            redirect_pos_login,
        )

        if not request.user.is_authenticated:
            return redirect_pos_login(request)
        if not is_staff_interno(request.user):
            return ocultar_sistema_interno(request)
        allowed_roles = [request.user.ROLE_ADMIN, request.user.ROLE_VENDEDOR, request.user.ROLE_TECNICO]
        if request.user.rol not in allowed_roles and not request.user.is_superuser:
            return ocultar_sistema_interno(request)
        return view_func(request, *args, **kwargs)
    return _wrapped_view

@reportes_required
def dashboard_reportes(request):
    """
    Main reports dashboard displaying sales KPIs, product rankings, and maintenance alerts.
    """
    now = timezone.now()
    
    # 1. Sales Calculations (Only PAID orders)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - datetime.timedelta(days=now.weekday())
    month_start = today_start.replace(day=1)
    
    ventas_hoy = Pedido.objects.filter(
        estado__in=Pedido.ESTADOS_CONCRETADOS,
        fecha_pedido__gte=today_start
    ).aggregate(total_sum=Sum('total'))['total_sum'] or Decimal('0.00')
    
    ventas_semana = Pedido.objects.filter(
        estado__in=Pedido.ESTADOS_CONCRETADOS,
        fecha_pedido__gte=week_start
    ).aggregate(total_sum=Sum('total'))['total_sum'] or Decimal('0.00')
    
    ventas_mes = Pedido.objects.filter(
        estado__in=Pedido.ESTADOS_CONCRETADOS,
        fecha_pedido__gte=month_start
    ).aggregate(total_sum=Sum('total'))['total_sum'] or Decimal('0.00')
    
    # Sales count
    count_hoy = Pedido.objects.filter(estado__in=Pedido.ESTADOS_CONCRETADOS, fecha_pedido__gte=today_start).count()
    count_mes = Pedido.objects.filter(estado__in=Pedido.ESTADOS_CONCRETADOS, fecha_pedido__gte=month_start).count()
    
    # 2. Top Selling Products
    productos_mas_vendidos = DetallePedido.objects.filter(
        pedido__estado__in=Pedido.ESTADOS_CONCRETADOS
    ).values(
        'producto__codigo_articulo', 'producto__nombre', 'producto__categoria__nombre'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        recaudacion_total=Sum('subtotal')
    ).order_by('-cantidad_vendida')[:5]

    # 3. Equipments Pending Maintenance (Hours >= Next Target & Active)
    equipos_alerta = EquipoRegistrado.objects.filter(
        horas_uso_actuales__gte=F('horas_proximo_mantenimiento'),
        estado=EquipoRegistrado.ESTADO_ACTIVO
    ).select_related('cliente', 'producto').order_by('-horas_uso_actuales')
    
    # 4. Quotes pending response (Draft or Sent)
    cotizaciones_pendientes = Cotizacion.objects.filter(
        estado__in=[Cotizacion.ESTADO_BORRADOR, Cotizacion.ESTADO_ENVIADA]
    ).select_related('cliente').order_by('-fecha_creacion')

    context = {
        'ventas_hoy': ventas_hoy,
        'ventas_semana': ventas_semana,
        'ventas_mes': ventas_mes,
        'count_hoy': count_hoy,
        'count_mes': count_mes,
        'productos_mas_vendidos': productos_mas_vendidos,
        'equipos_alerta': equipos_alerta,
        'cotizaciones_pendientes': cotizaciones_pendientes,
    }
    return render(request, 'reportes/dashboard.html', context)

@reportes_required
def exportar_reportes_excel(request):
    """
    Generate and stream an Excel workbook containing sales, products, and equipment alerts.
    """
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Style definitions
    font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    font_bold = Font(name='Arial', size=10, bold=True)
    fill_header = PatternFill(start_color='008B8B', end_color='008B8B', fill_type='solid') # Makita Teal
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ==========================================
    # SHEET 1: Ventas del Mes
    # ==========================================
    ws1 = wb.active
    ws1.title = "Ventas del Mes"
    ws1.views.sheetView[0].showGridLines = True
    
    # Headers
    headers1 = ["Fecha/Hora", "Nº Pedido", "Cliente", "Subtotal", "IGV", "Total", "Canal de Venta"]
    ws1.append(headers1)
    
    pedidos = Pedido.objects.filter(
        estado__in=Pedido.ESTADOS_CONCRETADOS,
        fecha_pedido__gte=month_start
    ).select_related('cliente').order_by('-fecha_pedido')
    
    for p in pedidos:
        ws1.append([
            p.fecha_pedido.strftime("%d/%m/%Y %H:%M"),
            p.numero_pedido,
            p.cliente.nombre_completo,
            float(p.subtotal),
            float(p.igv),
            float(p.total),
            p.get_canal_display()
        ])
        
    # Formatting sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        
    for cell in ws1[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in [4, 5, 6]:
                cell.number_format = 'S/. #,##0.00'
                cell.alignment = align_right
            elif cell.column in [1, 2, 7]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # ==========================================
    # SHEET 2: Productos Más Vendidos
    # ==========================================
    ws2 = wb.create_sheet(title="Productos Más Vendidos")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Código Artículo", "Producto / Modelo", "Categoría", "Cantidad Vendida", "Recaudación Total"]
    ws2.append(headers2)
    
    productos_vendidos = DetallePedido.objects.filter(
        pedido__estado__in=Pedido.ESTADOS_CONCRETADOS
    ).values(
        'producto__codigo_articulo', 'producto__nombre', 'producto__categoria__nombre'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        recaudacion_total=Sum('subtotal')
    ).order_by('-cantidad_vendida')
    
    for item in productos_vendidos:
        ws2.append([
            item['producto__codigo_articulo'],
            item['producto__nombre'],
            item['producto__categoria__nombre'] or "Sin Categoría",
            item['cantidad_vendida'],
            float(item['recaudacion_total'])
        ])
        
    # Formatting sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        
    for cell in ws2[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column == 5:
                cell.number_format = 'S/. #,##0.00'
                cell.alignment = align_right
            elif cell.column in [1, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # ==========================================
    # SHEET 3: Alertas Mantenimiento
    # ==========================================
    ws3 = wb.create_sheet(title="Alertas de Mantenimiento")
    ws3.views.sheetView[0].showGridLines = True
    
    headers3 = ["Nº Serie", "Modelo / Herramienta", "Cliente", "Horas Uso Actuales", "Horas Próximo Mantenimiento", "Diferencia Horas"]
    ws3.append(headers3)
    
    alertas = EquipoRegistrado.objects.filter(
        horas_uso_actuales__gte=F('horas_proximo_mantenimiento'),
        estado=EquipoRegistrado.ESTADO_ACTIVO
    ).select_related('cliente', 'producto').order_by('-horas_uso_actuales')
    
    for eq in alertas:
        modelo = eq.producto.codigo_articulo if eq.producto else eq.modelo_alternativo
        diff = eq.horas_uso_actuales - eq.horas_proximo_mantenimiento
        ws3.append([
            eq.numero_serie,
            modelo,
            eq.cliente.nombre_completo,
            eq.horas_uso_actuales,
            eq.horas_proximo_mantenimiento,
            diff
        ])
        
    # Formatting sheet 3
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        
    for cell in ws3[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in [4, 5, 6]:
                cell.alignment = align_right
            elif cell.column in [1, 2]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Stream workbook back
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_general_{now.strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    
    return response
