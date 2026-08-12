"""Exportación de OT y formulario de garantía desde plantillas Excel oficiales."""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook

TEMPLATES_DIR = Path(settings.BASE_DIR) / 'resources' / 'OT-612'
OT_TEMPLATE = TEMPLATES_DIR / 'OT 612.xlsx'
GARANTIA_TEMPLATE = TEMPLATES_DIR / 'FORMATO DE GARANTIA - 612.xlsx'


def _negocio():
    return getattr(settings, 'NEGOCIO', {}) or {}


def _parts_fecha(d):
    if not d:
        return '', '', ''
    if isinstance(d, datetime):
        d = d.date()
    return str(d.day), f'{d.month:02d}', str(d.year)[-2:]


def _sta_ruc():
    # RUC del STA en el formato Makita de ejemplo (persona/servicio técnico)
    return getattr(settings, 'STA_RUC_TECNICO', '10295867782')


def _sta_nombre():
    return getattr(settings, 'STA_NOMBRE_TECNICO', 'QUISPE PARI FREDY')


def exportar_ot_excel(mantenimiento) -> bytes:
    """Rellena la plantilla OT 612.xlsx y devuelve bytes del .xlsx."""
    equipo = mantenimiento.equipo
    cliente = equipo.cliente
    buf = io.BytesIO()
    # Copiar a memoria vía temp workbook
    wb = load_workbook(OT_TEMPLATE)
    ws = wb.active

    ot_num = mantenimiento.numero_ot_solo
    ws['AF2'] = int(ot_num) if ot_num.isdigit() else ot_num

    # Cliente (lado izquierdo)
    ws['G5'] = cliente.nombre_completo
    ws['G6'] = (cliente.direccion or '')[:80]
    ws['G7'] = cliente.telefono or ''
    ws['G8'] = mantenimiento.atencion_sr or cliente.nombre_completo
    # RUC label area — if there's a cell for RUC near ATT
    # Model / serie / descripción (zona derecha)
    ws['AB6'] = equipo.modelo_display
    ws['AH6'] = equipo.numero_serie
    ws['AB7'] = equipo.descripcion_display[:60]

    fr = mantenimiento.fecha_recepcion or (
        mantenimiento.fecha_ingreso.date() if mantenimiento.fecha_ingreso else None
    )
    d, m, y = _parts_fecha(fr)
    if d:
        ws['AF10'] = int(d) if d.isdigit() else d
        ws['AH10'] = m
        ws['AJ10'] = y
        ws['J68'] = int(d) if str(d).isdigit() else d
        ws['L68'] = int(m) if str(m).isdigit() else m
        ws['N68'] = int(y) if str(y).isdigit() else y

    fc = mantenimiento.fecha_compra or equipo.fecha_compra
    d2, m2, y2 = _parts_fecha(fc)
    if d2:
        ws['AF23'] = int(d2) if str(d2).isdigit() else d2
        ws['AH23'] = m2
        ws['AJ23'] = y2

    boleta = mantenimiento.boleta_factura or equipo.boleta_factura or ''
    ws['AB25'] = boleta
    ws['AB26'] = mantenimiento.distribuidor or equipo.distribuidor or (
        'El Charly' if equipo.vendido_por_nosotros else ''
    )

    if mantenimiento.tipo == mantenimiento.TIPO_GARANTIA:
        ws['AD17'] = 'X'

    tecnico_nombre = ''
    if mantenimiento.tecnico_id:
        tecnico_nombre = (
            mantenimiento.tecnico.get_full_name() or mantenimiento.tecnico.username
        )
    if tecnico_nombre:
        ws['AC29'] = tecnico_nombre
        ws['AA43'] = tecnico_nombre
        ws['AA54'] = tecnico_nombre

    # Observaciones / accesorios / informe
    obs = mantenimiento.observaciones or mantenimiento.diagnostico or ''
    if obs:
        ws['B34'] = obs[:500]
    if mantenimiento.accesorios:
        ws['B37'] = mantenimiento.accesorios[:500]
    informe = mantenimiento.informe_tecnico or mantenimiento.causa or ''
    if informe:
        ws['B41'] = informe[:2000]

    # Líneas de códigos (filas 14+)
    lineas = list(mantenimiento.lineas.all()[:8])
    # Clear sample row then fill
    for i, linea in enumerate(lineas):
        row = 14 + i
        ws.cell(row, 2).value = linea.cantidad  # B
        ws.cell(row, 5).value = linea.codigo  # E
        ws.cell(row, 9).value = linea.descripcion  # I

    wb.save(buf)
    wb.close()
    return buf.getvalue()


def exportar_garantia_excel(mantenimiento) -> bytes:
    """Rellena FORMATO DE GARANTIA - 612.xlsx."""
    equipo = mantenimiento.equipo
    cliente = equipo.cliente
    negocio = _negocio()
    buf = io.BytesIO()
    wb = load_workbook(GARANTIA_TEMPLATE)
    ws = wb['FORMULARIO_GARANTIA']

    # Named-range style cells (direct addresses)
    ws['D27'] = mantenimiento.autorizacion_mpe or None
    ws['G27'] = int(mantenimiento.numero_ot_solo) if mantenimiento.numero_ot_solo.isdigit() else mantenimiento.numero_ot_solo
    ws['D29'] = _sta_ruc()
    # D30+ suelen ser VLOOKUP desde BASE_STA; forzar nombre STA si vacío
    ws['D36'] = cliente.dni_ruc
    ws['D37'] = cliente.nombre_completo
    ws['D38'] = (cliente.direccion or '')[:120]
    ws['D39'] = getattr(cliente, 'ciudad', None) or negocio.get('ciudad', 'Arequipa')
    ws['D40'] = cliente.telefono or ''

    fr = mantenimiento.fecha_recepcion or (
        mantenimiento.fecha_ingreso.date() if mantenimiento.fecha_ingreso else None
    )
    fc = mantenimiento.fecha_compra or equipo.fecha_compra
    if fr:
        ws['G36'] = fr
    if fc:
        ws['G37'] = fc
    ws['G38'] = mantenimiento.boleta_factura or equipo.boleta_factura or ''

    ws['D44'] = equipo.modelo_display
    ws['G44'] = equipo.numero_serie

    ws['D57'] = (mantenimiento.diagnostico or '')[:200]
    ws['D58'] = (mantenimiento.causa or '')[:200]
    sustento = mantenimiento.informe_tecnico or mantenimiento.causa or ''
    # Sustento líneas C61+
    chunks = _chunk_text(sustento, 90, 8)
    for i, chunk in enumerate(chunks):
        ws.cell(61 + i, 3).value = chunk

    if mantenimiento.nombre_mpe:
        ws['D70'] = mantenimiento.nombre_mpe
    if mantenimiento.fecha_aprobacion_mpe:
        ws['G71'] = mantenimiento.fecha_aprobacion_mpe
    if mantenimiento.comentario_mpe:
        ws['D72'] = mantenimiento.comentario_mpe[:120]

    # Códigos C48:E54 (C47 es header NO BORRAR / códigos empiezan 48 en ejemplo MPE)
    lineas = list(mantenimiento.lineas.all()[:7])
    for i, linea in enumerate(lineas):
        row = 48 + i
        ws.cell(row, 3).value = linea.codigo
        ws.cell(row, 4).value = linea.descripcion
        ws.cell(row, 5).value = linea.cantidad

    # También resumen superior D4-D16 (panel izquierdo auxiliar)
    ws['D4'] = cliente.dni_ruc
    ws['D5'] = cliente.nombre_completo
    ws['D6'] = ws['D39'].value
    ws['D7'] = mantenimiento.numero_ot_solo
    if fr:
        ws['D8'] = fr
    if fc:
        ws['D9'] = fc
    ws['D10'] = mantenimiento.boleta_factura or equipo.boleta_factura or ''
    ws['D11'] = mantenimiento.distribuidor or equipo.distribuidor or ''
    ws['D12'] = equipo.modelo_display
    ws['D13'] = equipo.numero_serie
    ws['D14'] = (mantenimiento.diagnostico or '')[:80]
    ws['D15'] = (mantenimiento.causa or '')[:80]
    ws['D16'] = (sustento or '')[:200]
    if mantenimiento.autorizacion_mpe:
        ws['D1'] = mantenimiento.autorizacion_mpe
    if mantenimiento.nombre_mpe:
        ws['D21'] = mantenimiento.nombre_mpe
    if mantenimiento.comentario_mpe:
        ws['D22'] = mantenimiento.comentario_mpe[:80]
    if mantenimiento.mano_obra_mpe is not None:
        ws['D24'] = float(mantenimiento.mano_obra_mpe)

    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _chunk_text(text: str, size: int, max_lines: int) -> list[str]:
    text = (text or '').strip()
    if not text:
        return []
    lines = []
    while text and len(lines) < max_lines:
        lines.append(text[:size])
        text = text[size:]
    return lines


def filename_ot(mantenimiento) -> str:
    return f'{mantenimiento.numero_ot_display}.xlsx'


def filename_garantia(mantenimiento) -> str:
    base = mantenimiento.autorizacion_mpe or mantenimiento.numero_ot_display
    modelo = mantenimiento.equipo.modelo_display.replace(' ', '')[:20]
    return f'{base} {modelo}.xlsx'
