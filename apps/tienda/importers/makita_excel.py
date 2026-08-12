"""
Importación de catálogo Makita desde Excel mensual (hoja «Lista de Productos»).
No elimina productos existentes: solo inserta o actualiza. Registra cambios en log.
Soporta ejecución en segundo plano con bloqueo temporal de venta.
"""
from __future__ import annotations

import logging
import re
import threading
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any

from django.core.files.base import ContentFile
from django.db import close_old_connections, transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from apps.tienda.models import Categoria, Producto

logger = logging.getLogger(__name__)

SHEET_NAME = 'Lista de Productos'
MONEY_Q = Decimal('0.01')


def _norm_header(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip().upper()
    text = re.sub(r'\s+', ' ', text)
    return text


def _map_headers(row: tuple) -> dict[str, int]:
    mapping = {}
    aliases = {
        'codigo': {'COD. ARTICULO', 'COD ARTICULO', 'CODIGO ARTICULO', 'CÓD. ARTICULO', 'CODIGO'},
        'nombre': {'DESCRIPCION', 'DESCRIPCIÓN', 'NOMBRE'},
        'familia_sap': {'FAMILIA SAP', 'FAMILIA'},
        'categoria_sap': {'CATEGORIA SAP', 'CATEGORÍA SAP'},
        'categoria': {'CATEGORIA', 'CATEGORÍA'},
        'status': {'STATUS', 'STATUS SAP'},
        'precio': {'LISTA GENERAL', 'PRECIO', 'PRECIO LISTA', 'LISTA'},
    }
    for idx, cell in enumerate(row):
        header = _norm_header(cell)
        if not header:
            continue
        for field, names in aliases.items():
            if header in names and field not in mapping:
                mapping[field] = idx
    return mapping


def _cell(row: tuple, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    text = str(value).strip().replace(',', '').replace('S/', '').replace('S/.', '')
    if not text:
        return None
    try:
        return Decimal(text).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def _norm_color_lista(value: str) -> str | None:
    """Normaliza el color/STATUS del Excel (Verde, Rojo, …)."""
    raw = (value or '').strip()
    if not raw:
        return None
    key = raw.casefold()
    known = {
        'verde': 'Verde',
        'amarillo': 'Amarillo',
        'rojo': 'Rojo',
        'morado': 'Morado',
        'azul': 'Azul',
        'violeta': 'Morado',
        'yellow': 'Amarillo',
        'green': 'Verde',
        'red': 'Rojo',
        'purple': 'Morado',
        'blue': 'Azul',
    }
    if key in known:
        return known[key]
    # A veces viene "STATUS VERDE" u otro texto con el color embebido
    for token, label in known.items():
        if token in key:
            return label
    return raw[:50]


def _tipo_from_familia(familia: str) -> str:
    fam = (familia or '').strip().upper()
    if 'REPUESTO' in fam:
        return Producto.TIPO_REPUESTO
    if 'ACCESORIO' in fam:
        return Producto.TIPO_ACCESORIO
    if 'HERRAMIENT' in fam or 'EQUIPO' in fam or 'PROMO' in fam:
        return Producto.TIPO_HERRAMIENTA
    return Producto.TIPO_ACCESORIO


def _extract_modelo(descripcion: str) -> str:
    if not descripcion:
        return ''
    match = re.search(r'/\s*([A-Z0-9][A-Z0-9\-]{2,})\s*$', descripcion.strip(), re.I)
    if match:
        return match.group(1)[:100]
    return ''


def _get_or_create_categoria(nombre: str | None) -> Categoria | None:
    nombre = (nombre or '').strip()
    if not nombre:
        return None
    slug = slugify(nombre)[:120] or 'categoria'
    cat = Categoria.objects.filter(slug=slug).first()
    if cat:
        return cat
    cat = Categoria.objects.filter(nombre__iexact=nombre).first()
    if cat:
        return cat
    return Categoria.objects.create(nombre=nombre[:100], slug=slug)


def find_lista_productos_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == SHEET_NAME.lower():
            return wb[name]
    return wb[wb.sheetnames[-1]]


def parse_makita_excel(file_obj) -> list[dict]:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = find_lista_productos_sheet(wb)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = _map_headers(tuple(header_row or ()))
        if 'codigo' not in headers or 'nombre' not in headers or 'precio' not in headers:
            raise ValueError(
                'La hoja «Lista de Productos» debe incluir columnas '
                'COD. ARTICULO, DESCRIPCION y LISTA GENERAL.'
            )

        items = []
        for row in rows_iter:
            if not row:
                continue
            codigo = str(_cell(row, headers['codigo']) or '').strip()
            if not codigo or codigo.upper() in {'COD. ARTICULO', 'NONE'}:
                continue
            nombre = str(_cell(row, headers['nombre']) or '').strip()
            if not nombre:
                continue
            precio = _to_decimal(_cell(row, headers.get('precio')))
            if precio is None:
                continue
            # Incluir precios «basura» tipo 9999999 / 999999999: el producto
            # debe existir en inventario interno aunque no sea vendible a ese monto.
            if precio < 0:
                continue
            if precio > Decimal('9999999999.99'):
                precio = Decimal('9999999999.99')
            familia = str(_cell(row, headers.get('familia_sap')) or '').strip()
            categoria_sap = str(_cell(row, headers.get('categoria_sap')) or '').strip()
            categoria = str(_cell(row, headers.get('categoria')) or familia or '').strip()
            # STATUS del Excel Makita suele ser el color de fila
            # (Verde/Amarillo/Rojo/Morado/Azul) = disponibilidad Lima.
            status_raw = str(_cell(row, headers.get('status')) or '').strip()
            status = _norm_color_lista(status_raw)  # puede ser None
            items.append({
                'codigo_articulo': codigo[:50],
                'nombre': nombre[:255],
                'familia_sap': (familia or 'SIN FAMILIA')[:100],
                'categoria_sap': (categoria_sap[:100] if categoria_sap else None),
                'categoria_nombre': (categoria[:100] if categoria else None),
                'status_sap': (status[:50] if status else None),
                'precio_venta': precio,
                'tipo': _tipo_from_familia(familia),
                'modelo': _extract_modelo(nombre) or None,
            })
        return items
    finally:
        wb.close()


def _bloquear_codigos(codigos: list[str]) -> int:
    if not codigos:
        return 0
    total = 0
    chunk = 500
    for i in range(0, len(codigos), chunk):
        parte = codigos[i:i + chunk]
        total += Producto.objects.filter(codigo_articulo__in=parte).exclude(tipo=Producto.TIPO_SERVICIO).update(venta_bloqueada=True)
    return total


def _desbloquear_codigos(codigos: list[str]) -> int:
    if not codigos:
        return 0
    total = 0
    chunk = 500
    for i in range(0, len(codigos), chunk):
        parte = codigos[i:i + chunk]
        total += Producto.objects.filter(codigo_articulo__in=parte).update(venta_bloqueada=False)
    return total


def _motivo_importacion(imp) -> str:
    nombre = getattr(imp, 'archivo_nombre', '') or 'lista Makita'
    return f'Actualización por importación Excel «{nombre}»'


def eliminar_archivo_excel(imp, *, motivo: str = '') -> bool:
    """
    Borra el Excel del disco y limpia el FileField.
    Se usa tras importación exitosa (los datos ya están en BD + log de cambios).
    """
    if not imp or not getattr(imp, 'archivo', None):
        return False
    try:
        name = imp.archivo.name
        if name:
            imp.archivo.delete(save=False)
        type(imp).objects.filter(pk=imp.pk).update(archivo='')
        logger.info(
            'Excel eliminado tras importación #%s%s',
            imp.pk,
            f' ({motivo})' if motivo else '',
        )
        return True
    except Exception:
        logger.exception('No se pudo eliminar Excel de importación #%s', getattr(imp, 'pk', '?'))
        return False


def limpiar_excels_importaciones_exitosas() -> int:
    """Elimina Excel de importaciones ya completadas sin errores de fila."""
    from apps.tienda.models import ImportacionCatalogo

    qs = ImportacionCatalogo.objects.filter(
        estado=ImportacionCatalogo.ESTADO_COMPLETADA,
        total_errores=0,
    ).exclude(archivo='').exclude(archivo__isnull=True)
    borrados = 0
    for imp in qs.iterator():
        if eliminar_archivo_excel(imp, motivo='limpieza post-éxito'):
            borrados += 1
    return borrados


def _totales_desde_log(importacion_id: int, total_filas: int) -> tuple[int, int, int, int]:
    """
    Recalcula nuevos / actualizados / sin_cambio / errores desde el log.
    Un SKU cuenta una sola vez (aunque tenga varios cambios: precio + nombre…).
    """
    from apps.tienda.models import LogCambioImportacion

    qs = LogCambioImportacion.objects.filter(importacion_id=importacion_id)
    nuevos = (
        qs.filter(tipo_cambio=LogCambioImportacion.TIPO_NUEVO)
        .values('codigo_articulo')
        .distinct()
        .count()
    )
    errores = (
        qs.filter(tipo_cambio=LogCambioImportacion.TIPO_ERROR)
        .values('codigo_articulo')
        .distinct()
        .count()
    )
    actualizados = (
        qs.exclude(
            tipo_cambio__in=[
                LogCambioImportacion.TIPO_NUEVO,
                LogCambioImportacion.TIPO_ERROR,
            ]
        )
        .values('codigo_articulo')
        .distinct()
        .count()
    )
    sin_cambio = max(0, total_filas - nuevos - actualizados - errores)
    return nuevos, actualizados, sin_cambio, errores


def _log_cambio(imp, *, codigo, tipo, campo, anterior='', nuevo='', detalle=''):
    from apps.tienda.models import LogCambioImportacion

    LogCambioImportacion.objects.create(
        importacion=imp,
        codigo_articulo=codigo,
        tipo_cambio=tipo,
        campo=campo or '',
        valor_anterior=str(anterior or '')[:2000],
        valor_nuevo=str(nuevo or '')[:2000],
        detalle=(detalle or '')[:2000],
    )


def _aplicar_item(imp, item: dict) -> str:
    """Aplica un ítem. Retorna: 'nuevo' | 'actualizado' | 'sin_cambio'."""
    from apps.tienda.models import LogCambioImportacion
    from apps.tienda.precios import con_igv

    codigo = item['codigo_articulo']
    categoria = _get_or_create_categoria(item['categoria_nombre'] or item['familia_sap'])
    existing = Producto.objects.filter(codigo_articulo=codigo).first()
    if existing and existing.tipo == Producto.TIPO_SERVICIO:
        return 'sin_cambio'
    motivo = _motivo_importacion(imp)

    if existing is None:
        Producto.objects.create(
            codigo_articulo=codigo,
            nombre=item['nombre'],
            familia_sap=item['familia_sap'],
            categoria_sap=item['categoria_sap'],
            status_sap=item['status_sap'],
            precio_venta=item['precio_venta'],
            tipo=item['tipo'],
            modelo=item['modelo'],
            categoria=categoria,
            marca='Makita',
            activo=True,
            mostrar_en_web=False,  # no publicar en web hasta selección manual
            stock=0,
            venta_bloqueada=False,
        )
        _log_cambio(
            imp,
            codigo=codigo,
            tipo=LogCambioImportacion.TIPO_NUEVO,
            campo='alta',
            nuevo=item['nombre'],
            detalle=(
                f'Producto nuevo en catálogo. Lista S/ {item["precio_venta"]} '
                f'(con IGV S/ {con_igv(item["precio_venta"])}). {motivo}.'
            ),
        )
        return 'nuevo'

    hubo_cambio = False

    if (existing.nombre or '') != (item['nombre'] or ''):
        ant, nue = existing.nombre or '', item['nombre'] or ''
        _log_cambio(
            imp,
            codigo=codigo,
            tipo=LogCambioImportacion.TIPO_NOMBRE,
            campo='nombre',
            anterior=ant,
            nuevo=nue,
            detalle=(
                f'Nombre/descripción cambió: «{ant[:80]}» → «{nue[:80]}». '
                f'Motivo: nueva DESCRIPCION en Excel. {motivo}.'
            ),
        )
        existing.nombre = item['nombre']
        existing.slug = slugify(f"{item['nombre']}-{codigo}")[:280]
        hubo_cambio = True

    if existing.precio_venta != item['precio_venta']:
        lista_ant, lista_nue = existing.precio_venta, item['precio_venta']
        sube = lista_nue > lista_ant
        tipo_cambio = (
            LogCambioImportacion.TIPO_PRECIO_SUBE
            if sube
            else LogCambioImportacion.TIPO_PRECIO_BAJA
        )
        sentido = 'subió' if sube else 'bajó'
        delta = abs(lista_nue - lista_ant)
        _log_cambio(
            imp,
            codigo=codigo,
            tipo=tipo_cambio,
            campo='precio_venta',
            anterior=str(lista_ant),
            nuevo=str(lista_nue),
            detalle=(
                f'Precio lista {sentido}: S/ {lista_ant} → S/ {lista_nue} '
                f'(Δ S/ {delta}). Con IGV: S/ {con_igv(lista_ant)} → S/ {con_igv(lista_nue)}. '
                f'Motivo: nueva LISTA GENERAL en Excel. {motivo}.'
            ),
        )
        existing.precio_venta = item['precio_venta']
        hubo_cambio = True

    # Color Lima / STATUS
    status_ant = (existing.status_sap or '').strip()
    status_nue = (item.get('status_sap') or '').strip()
    if status_ant != status_nue and (status_ant or status_nue):
        _log_cambio(
            imp,
            codigo=codigo,
            tipo=LogCambioImportacion.TIPO_STATUS,
            campo='status_sap',
            anterior=status_ant or '(vacío)',
            nuevo=status_nue or '(vacío)',
            detalle=(
                f'Color lista Lima: {status_ant or "—"} → {status_nue or "—"}. '
                f'Motivo: columna Status del Excel. {motivo}.'
            ),
        )
        hubo_cambio = True

    familia_ant = (existing.familia_sap or '').strip()
    familia_nue = (item.get('familia_sap') or '').strip()
    if familia_ant != familia_nue and familia_nue:
        _log_cambio(
            imp,
            codigo=codigo,
            tipo=LogCambioImportacion.TIPO_FAMILIA,
            campo='familia_sap',
            anterior=familia_ant or '(vacío)',
            nuevo=familia_nue,
            detalle=(
                f'Familia SAP: {familia_ant or "—"} → {familia_nue}. '
                f'Motivo: columna FAMILIA SAP del Excel. {motivo}.'
            ),
        )
        hubo_cambio = True

    existing.familia_sap = item['familia_sap']
    existing.categoria_sap = item['categoria_sap']
    existing.status_sap = item['status_sap']
    existing.tipo = item['tipo']
    if item['modelo']:
        existing.modelo = item['modelo']
    if categoria:
        existing.categoria = categoria

    if not existing.activo:
        existing.activo = True
        _log_cambio(
            imp,
            codigo=codigo,
            tipo=LogCambioImportacion.TIPO_REACTIVADO,
            campo='activo',
            anterior='False',
            nuevo='True',
            detalle=f'Producto reactivado (estaba inactivo). {motivo}.',
        )
        hubo_cambio = True

    # Liberar venta de este SKU al terminar su fila
    existing.venta_bloqueada = False
    existing.save()
    return 'actualizado' if hubo_cambio else 'sin_cambio'


def procesar_importacion(importacion_id: int) -> None:
    """Procesa una ImportacionCatalogo (hilo / management command)."""
    from apps.tienda.models import ImportacionCatalogo

    close_old_connections()
    codigos_bloqueados: list[str] = []

    try:
        try:
            imp = ImportacionCatalogo.objects.get(pk=importacion_id)
        except ImportacionCatalogo.DoesNotExist:
            return

        imp.estado = ImportacionCatalogo.ESTADO_PROCESANDO
        imp.mensaje_error = ''
        imp.save(update_fields=['estado', 'mensaje_error'])

        if not imp.archivo:
            raise ValueError('No hay archivo asociado a la importación.')

        with imp.archivo.open('rb') as fh:
            items = parse_makita_excel(fh)

        from apps.tienda.models import LogCambioImportacion

        # Si se reanuda, borrar logs previos para que el resumen coincida con esta corrida
        LogCambioImportacion.objects.filter(importacion=imp).delete()

        imp.total_filas = len(items)
        imp.total_procesadas = 0
        imp.total_nuevos = 0
        imp.total_actualizados = 0
        imp.total_sin_cambio = 0
        imp.total_errores = 0
        imp.save(update_fields=[
            'total_filas', 'total_procesadas', 'total_nuevos',
            'total_actualizados', 'total_sin_cambio', 'total_errores',
        ])

        codigos = [i['codigo_articulo'] for i in items]
        existentes = []
        for i in range(0, len(codigos), 500):
            existentes.extend(
                Producto.objects.filter(codigo_articulo__in=codigos[i:i + 500])
                .values_list('codigo_articulo', flat=True)
            )
        codigos_bloqueados = list(existentes)
        _bloquear_codigos(codigos_bloqueados)

        nuevos = actualizados = sin_cambio = errores = 0
        for idx, item in enumerate(items, start=1):
            try:
                with transaction.atomic():
                    resultado = _aplicar_item(imp, item)
                if resultado == 'nuevo':
                    nuevos += 1
                elif resultado == 'actualizado':
                    actualizados += 1
                else:
                    sin_cambio += 1
            except Exception as row_exc:
                errores += 1
                logger.exception('Fila %s SKU %s', idx, item.get('codigo_articulo'))
                LogCambioImportacion.objects.create(
                    importacion=imp,
                    codigo_articulo=item.get('codigo_articulo', '')[:50],
                    tipo_cambio=LogCambioImportacion.TIPO_ERROR,
                    campo='fila',
                    detalle=str(row_exc)[:255],
                    valor_nuevo=item.get('nombre', '')[:500],
                )
                # Liberar bloqueo de este SKU si falló
                Producto.objects.filter(
                    codigo_articulo=item.get('codigo_articulo')
                ).update(venta_bloqueada=False)

            if idx % 25 == 0 or idx == len(items):
                ImportacionCatalogo.objects.filter(pk=imp.pk).update(
                    total_procesadas=idx,
                    total_nuevos=nuevos,
                    total_actualizados=actualizados,
                    total_sin_cambio=sin_cambio,
                    total_errores=errores,
                )

        # Totales finales alineados con el log (fuente de verdad)
        nuevos, actualizados, sin_cambio, errores = _totales_desde_log(imp.pk, len(items))

        ImportacionCatalogo.objects.filter(pk=imp.pk).update(
            estado=ImportacionCatalogo.ESTADO_COMPLETADA,
            total_procesadas=len(items),
            total_nuevos=nuevos,
            total_actualizados=actualizados,
            total_sin_cambio=sin_cambio,
            total_errores=errores,
            fecha_fin=timezone.now(),
            mensaje_error='',
        )
        _desbloquear_codigos(codigos_bloqueados)

        # Éxito sin errores de fila → borrar Excel (data ya en BD + log).
        # Si hubo errores por fila, se conserva para revisar / reanudar.
        if errores == 0:
            imp.refresh_from_db(fields=['archivo'])
            eliminar_archivo_excel(imp, motivo='completada sin errores')

    except Exception as exc:
        logger.exception('Error en importación %s', importacion_id)
        try:
            from apps.tienda.models import ImportacionCatalogo
            ImportacionCatalogo.objects.filter(pk=importacion_id).update(
                estado=ImportacionCatalogo.ESTADO_ERROR,
                mensaje_error=str(exc)[:2000],
                fecha_fin=timezone.now(),
            )
            # En error el Excel se mantiene para reintentar / diagnosticar
        except Exception:
            pass
        _desbloquear_codigos(codigos_bloqueados)
    finally:
        close_old_connections()


def iniciar_importacion_en_background(importacion_id: int) -> None:
    """
    Lanza un proceso separado (sobrevive a reinicios del runserver / hilos daemon).
    Fallback a hilo solo si el subprocess falla al arrancar.
    """
    import subprocess
    import sys
    from pathlib import Path

    from django.conf import settings

    manage = str(Path(settings.BASE_DIR) / 'manage.py')
    cmd = [sys.executable, manage, 'procesar_importacion_catalogo', str(importacion_id)]
    try:
        subprocess.Popen(
            cmd,
            cwd=str(settings.BASE_DIR),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
        logger.info('Importación #%s lanzada en subprocess', importacion_id)
        return
    except Exception:
        logger.exception('No se pudo lanzar subprocess; usando hilo daemon')

    thread = threading.Thread(
        target=procesar_importacion,
        args=(importacion_id,),
        name=f'import-catalogo-{importacion_id}',
        daemon=True,
    )
    thread.start()


def importar_catalogo_makita(
    file_obj,
    archivo_nombre: str = '',
    tipo_archivo: str = 'auto',
    usuario=None,
    en_background: bool = True,
):
    """
    Guarda el Excel y procesa (por defecto en segundo plano).
    Retorna la instancia ImportacionCatalogo.
    """
    from apps.tienda.models import ImportacionCatalogo

    nombre = archivo_nombre or getattr(file_obj, 'name', 'catalogo.xlsx')
    data = file_obj.read() if hasattr(file_obj, 'read') else file_obj
    if hasattr(file_obj, 'seek'):
        try:
            file_obj.seek(0)
        except Exception:
            pass

    if not data:
        raise ValueError('El archivo Excel está vacío.')

    imp = ImportacionCatalogo(
        archivo_nombre=nombre[:255],
        tipo_archivo=tipo_archivo or ImportacionCatalogo.TIPO_AUTO,
        usuario=usuario if getattr(usuario, 'is_authenticated', False) else None,
        estado=ImportacionCatalogo.ESTADO_PENDIENTE,
    )
    # Guardar archivo de inmediato (antes fallaba y quedaba pendiente sin archivo)
    safe_name = nombre.replace('\\', '_').replace('/', '_')[:200]
    if not safe_name.lower().endswith(('.xlsx', '.xlsm')):
        safe_name = f'{safe_name}.xlsx'
    imp.save()  # pk primero
    imp.archivo.save(safe_name, ContentFile(data), save=True)
    if not imp.archivo:
        raise ValueError('No se pudo guardar el archivo de importación en disco.')

    if en_background:
        iniciar_importacion_en_background(imp.pk)
    else:
        procesar_importacion(imp.pk)
        imp.refresh_from_db()

    return imp


# Alias legacy
def importar_catalogo(*args, **kwargs):
    return importar_catalogo_makita(*args, **kwargs)
